from io import BytesIO
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Результат")

        workbook = writer.book
        worksheet = writer.sheets["Результат"]

        wrap_format = workbook.add_format({
            "text_wrap": True,
            "valign": "top"
        })

        default_format = workbook.add_format({
            "valign": "top"
        })

        for col_num, column_name in enumerate(df.columns):
            if column_name in ["Товары заявки", "Список товаров", "Комментарий", "Адрес доставки"]:
                worksheet.set_column(col_num, col_num, 60, wrap_format)
            else:
                worksheet.set_column(col_num, col_num, 22, default_format)

    return output.getvalue()


def format_quantity(value) -> str:
    if pd.isna(value):
        return ""

    value_str = str(value).replace("\xa0", " ").strip()

    if value_str == "":
        return ""

    value_str = value_str.replace(",", ".")

    try:
        number = float(value_str)

        if number.is_integer():
            return str(int(number))

        return str(number).rstrip("0").rstrip(".")

    except ValueError:
        return value_str


def clean_phone(value):
    if pd.isna(value):
        return pd.NA

    value_str = str(value).replace("\xa0", " ").strip()

    if value_str == "":
        return pd.NA

    value_str = value_str.replace(" ", "")

    try:
        number = float(value_str)

        if number.is_integer():
            return str(int(number))

    except ValueError:
        pass

    if value_str.endswith(".0"):
        value_str = value_str[:-2]

    return value_str


def join_products_with_quantity(group: pd.DataFrame) -> str:
    product_lines = []

    for _, row in group.iterrows():
        product = row.get("Список товаров")
        quantity = row.get("Кол-во товара")

        if pd.isna(product):
            continue

        product_text = str(product).replace("\xa0", " ").strip()

        if product_text == "":
            continue

        quantity_text = format_quantity(quantity)

        if quantity_text != "":
            product_lines.append(f"{product_text} - {quantity_text}шт.")
        else:
            product_lines.append(product_text)

    return "\n".join(product_lines)


def prepare_drivers_file(vod_file) -> pd.DataFrame:
    df_vod = pd.read_excel(vod_file)

    # Чистим названия столбцов: убираем переносы, неразрывные пробелы и лишние пробелы
    df_vod.columns = (
        df_vod.columns
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.replace("\n", " ", regex=False)
        .str.replace("\r", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    required_columns = ["ФИО водителя", "Номер заявки"]

    missing_columns = [
        col for col in required_columns
        if col not in df_vod.columns
    ]

    if missing_columns:
        found_columns = ", ".join(map(str, df_vod.columns))

        raise ValueError(
            "Файл с водителями должен содержать ровно два столбца: "
            "«ФИО водителя» и «Номер заявки». "
            f"Не найдены столбцы: {', '.join(missing_columns)}. "
            f"Найденные столбцы в файле: {found_columns}"
        )

    # Берём только нужные два столбца.
    # Порядок в Excel не важен: может быть сначала ФИО, потом номер заявки.
    df_vod = df_vod[["Номер заявки", "ФИО водителя"]].copy()

    df_vod["Номер заявки"] = (
        df_vod["Номер заявки"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", "", regex=True)
        .str.strip()
    )

    df_vod["ФИО водителя"] = (
        df_vod["ФИО водителя"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    df_vod = df_vod.dropna(subset=["Номер заявки"])
    df_vod = df_vod[df_vod["Номер заявки"] != ""]

    df_vod = df_vod.drop_duplicates(
        subset=["Номер заявки"],
        keep="first"
    ).reset_index(drop=True)

    df_vod = df_vod.rename(columns={
        "ФИО водителя": "Водитель"
    })

    return df_vod


def make_products_df(df: pd.DataFrame, group_columns: list[str]) -> pd.DataFrame:
    rows = []

    for group_key, group in df.groupby(group_columns, dropna=False, sort=False):
        if not isinstance(group_key, tuple):
            group_key = (group_key,)

        row = dict(zip(group_columns, group_key))
        row["Товары заявки"] = join_products_with_quantity(group)
        rows.append(row)

    return pd.DataFrame(rows)


def make_short_df(grouped_df: pd.DataFrame) -> pd.DataFrame:
    if "Водитель" in grouped_df.columns:
        short_columns = [
            "Дата доставки",
            "Номер заявки",
            "Водитель",
            "Филиал",
            "Вес заказа",
            "Адрес доставки",
            "Стоимость заказа, руб.",
            "Зона"
        ]
    else:
        short_columns = [
            "Дата доставки",
            "Номер заявки",
            "Филиал",
            "Вес заказа",
            "Адрес доставки",
            "Стоимость заказа, руб.",
            "Зона"
        ]

    return grouped_df[short_columns].copy()



# --------------------------------------------------
# Логика зон для вкладки «Работа с реестром»
# --------------------------------------------------

ZONE_0_PLACES = [
    "Янино-1", "Кудрово", "Парголово", "Юкки", "Новогорелово",
    "Новоселье", "Левашово", "Мурино", "Бугры", "Новое Девяткино",
    "Ковалёво", "Лисий нос", "Лаврики",
]

ZONE_1_PLACES = [
    "Пушкин", "Лесколово", "Сертолово", "Янино-2", "Сестрорецк",
    "Петергоф", "Ломоносов", "Кронштадт", "Колтуши", "Металлострой",
    "Колпино", "Красное село", "Павловск", "Коммунар", "Виллози",
    "Аннино", "Песочный", "Агалатово", "Вартемяги", "Лупполово",
    "Мистолово", "Кузьмолово", "Энколово", "Капитолово", "Всеволожск",
    "Отрадное", "Никольское", "Кипень", "Красный Бор", "Ропша", "Низино",
]

ZONE_2_PLACES = [
    "Гатчина", "Шлиссельбург", "Кировск", "посёлок имени Морозова",
    "Ульяновка", "Новый свет", "Гостилицы", "Сельцо",
]

ZONE_3_PLACES = ["Мга"]

ZONE_0_STREETS = [
    "Улица Савушкина",
    "Комендантский проспект",
    "Богатырский проспект",
    "Шуваловский проспект",
    "Планерная улица",
    "Парашютная улица",
    "Проспект Испытателей",
    "Проспект Королёва",
    "Проспект Авиаконструкторов",
    "Улица Оптиков",
    "Мебельная улица",
    "Камышовая улица",
    "Улица Маршала Новикова",
    "Главная улица",
    "Дибуновская улица",
    "Приморское шоссе",
    "Петровская аллея",
    "Дорога в Каменку",
    "Большой Сампсониевский проспект",
    "Проспект Энгельса",
    "Суздальский проспект",
    "Проспект Культуры",
    "Лесной проспект",
    "Тихорецкий проспект",
    "Проспект Тореза",
    "Светлановский проспект",
    "Проспект Луначарского",
    "Проспект Просвещения",
    "Северный проспект",
    "Проспект Художников",
    "Удельный проспект",
    "Политехническая улица",
    "Улица Есенина",
    "Улица Асафьева",
    "Улица Сантьяго-де-Куба",
    "Проспект Стачек",
    "Ленинский проспект",
    "Проспект Маршала Жукова",
    "Проспект Ветеранов",
    "Проспект Народного Ополчения",
    "Краснопутиловская улица",
    "Дачный проспект",
    "Улица Стойкости",
    "Улица Маршала Казакова",
    "Улица Генерала Симоняка",
    "Улица Зины Портновой",
    "Улица Лёни Голикова",
    "Улица Васи Алексеева",
    "Улица Белоусова",
    "Улица Гладкова",
    "Улица Трефолева",
    "Улица Косинова",
    "Улица Возрождения",
    "Улица Зайцева",
    "Улица Подводника Кузьмина",
    "Проспект Большевиков",
    "Дальневосточный проспект",
    "Искровский проспект",
    "Проспект Обуховской Обороны",
    "Шлиссельбургский проспект",
    "Российский проспект",
    "Проспект Пятилеток",
    "Улица Дыбенко",
    "Улица Коллонтай",
    "Улица Крыленко",
    "Улица Тельмана",
    "Улица Народная",
    "Улица Латышских Стрелков",
    "Улица Бабушкина",
    "Ивановская улица",
    "Октябрьская набережная",
    "Набережная Обводного канала",
    "Улица Седова",
    "Улица Новосёлов",
    "Улица Подвойского",
    "Невский проспект",
    "Литейный проспект",
    "Владимирский проспект",
    "Загородный проспект",
    "Суворовский проспект",
    "Лиговский проспект",
    "Садовая улица",
    "Большая Морская улица",
    "Малая Морская улица",
    "Большая Конюшенная улица",
    "Малая Конюшенная улица",
    "Итальянская улица",
    "Гороховая улица",
    "Миллионная улица",
    "Шпалерная улица",
    "Улица Восстания",
    "Улица Марата",
    "Улица Некрасова",
    "Улица Рубинштейна",
    "Набережная реки Фонтанки",
    "Вознесенский проспект",
    "Английский проспект",
    "Большая Подьяческая улица",
    "улица Декабристов",
    "Большой проспект Васильевского острова",
    "Университетская набережная",
    "6-я линия Васильевского острова",
    "7-я линия Васильевского острова",
    "улица Шевченко",
    "Большой проспект Петроградской стороны",
    "Каменноостровский проспект",
    "Кронверкский проспект",
    "Бармалеева улица",
    "Кондратьевский проспект",
    "Гражданский проспект",
    "улица Руставели",
    "улица Верности",
    "Большеохтинский проспект",
    "Среднеохтинский проспект",
    "проспект Шаумяна",
    "Заневский проспект",
    "Новоизмайловский проспект",
    "улица Гастелло",
    "Бассейная улица",
    "проспект Славы",
    "Бухарестская улица",
    "Дунайский проспект",
    "улица Белы Куна",
    "Пискарёвский проспект",
]


def normalize_zone_address(value, remove_street_type=False) -> str:
    if pd.isna(value):
        return ""

    text = str(value).replace("\xa0", " ").lower().replace("ё", "е")

    if remove_street_type:
        text = re.sub(
            r"(?<!\w)(?:улица|ул\.?|проспект|пр-?т|пр-?кт|пр\.?)(?!\w)",
            " ",
            text,
            flags=re.IGNORECASE,
        )

    text = re.sub(r"[^0-9a-zа-я]+", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def zone_phrase_in_address(address_text: str, phrase: str, remove_street_type=False) -> bool:
    normalized_address = normalize_zone_address(address_text, remove_street_type)
    normalized_phrase = normalize_zone_address(phrase, remove_street_type)

    if normalized_address == "" or normalized_phrase == "":
        return False

    pattern = rf"(?<!\w){re.escape(normalized_phrase)}(?!\w)"
    return re.search(pattern, normalized_address, flags=re.IGNORECASE) is not None


def determine_delivery_zone(address) -> int:
    # Населённый пункт имеет приоритет над улицей.
    for place in ZONE_0_PLACES:
        if zone_phrase_in_address(address, place):
            return 0

    for place in ZONE_1_PLACES:
        if zone_phrase_in_address(address, place):
            return 1

    for place in ZONE_2_PLACES:
        if zone_phrase_in_address(address, place):
            return 2

    for place in ZONE_3_PLACES:
        if zone_phrase_in_address(address, place):
            return 3

    # Для улиц игнорируем написание «улица/ул.» и «проспект/пр./пр-т»,
    # а также положение обозначения до или после названия.
    for street in ZONE_0_STREETS:
        if zone_phrase_in_address(address, street, remove_street_type=True):
            return 0

    # Санкт-Петербург без совпадения с улицами зоны 0 и любые прочие адреса.
    return 8


def process_delivery_file(main_file, drivers_file=None, original_filename: str = "file.xlsx"):
    df = pd.read_excel(
        main_file,
        sheet_name="TDSheet"
    )

    # --------------------------------------------------
    # 1. Добавляем филиал
    # --------------------------------------------------

    current_filial = None
    filial_values = []

    for i in range(len(df)):
        first_col_value = df.iloc[i, 0]

        if df.shape[1] > 4:
            fifth_col_value = df.iloc[i, 4]
        else:
            fifth_col_value = None

        if pd.notna(first_col_value) and "Филиал" in str(first_col_value):
            if pd.notna(fifth_col_value) and str(fifth_col_value).strip() != "":
                current_filial = str(fifth_col_value).strip()

        filial_values.append(current_filial)

    df["Филиал"] = filial_values

    # --------------------------------------------------
    # 2. ВАЖНО: не удаляем пустые столбцы
    # --------------------------------------------------
    # Раньше здесь был dropna(axis=1, how="all").
    # Теперь его нет, чтобы нужные, но пустые столбцы не пропадали.

    df = df.replace(r"^\s*$", pd.NA, regex=True)

    # --------------------------------------------------
    # 3. Удаляем лишние строки
    # --------------------------------------------------

    first_col = df.iloc[:, 0].astype(str)

    mask_delete = (
        first_col.str.contains("Филиал:", na=False) |
        first_col.str.contains("Список заявок на доставку", na=False) |
        first_col.str.contains("Доставочная организация:", na=False) |
        first_col.str.contains("Дата:", na=False) |
        first_col.str.contains("№", na=False)
    )

    mask_itogo = df.astype("string").apply(
        lambda col: col.str.contains("ИТОГО", case=False, na=False)
    ).any(axis=1)

    mask_delete = mask_delete | mask_itogo

    cols_except_filial = [col for col in df.columns if col != "Филиал"]
    mask_empty_except_filial = df[cols_except_filial].isna().all(axis=1)

    mask_delete = mask_delete | mask_empty_except_filial

    df = df[~mask_delete].reset_index(drop=True)

    # --------------------------------------------------
    # 4. Забираем нужные исходные столбцы
    # --------------------------------------------------
    # Если нужный столбец оказался полностью пустым и pandas его не прочитал,
    # создаём его пустым, чтобы на выходе структура всегда была одинаковой.

    source_columns = [
        "Unnamed: 3",
        "Unnamed: 5",
        "Unnamed: 8",
        "Unnamed: 11",
        "Unnamed: 13",
        "Unnamed: 20",
        "Unnamed: 21",
        "Unnamed: 23",
        "Unnamed: 25",
        "Unnamed: 27",
        "Unnamed: 29",
        "Unnamed: 35",
        "Unnamed: 51",
        "Филиал"
    ]

    for col in source_columns:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[source_columns].copy()

    # --------------------------------------------------
    # 5. Заполняем пропуски вниз, кроме комментария
    # --------------------------------------------------

    exclude_col = "Unnamed: 51"

    cols_to_fill = [col for col in df.columns if col != exclude_col]

    df[cols_to_fill] = df[cols_to_fill].replace(r"^\s*$", pd.NA, regex=True)
    df[cols_to_fill] = df[cols_to_fill].ffill()

    # --------------------------------------------------
    # 6. Переименовываем столбцы
    # --------------------------------------------------

    df = df.rename(columns={
        "Unnamed: 3": "Номер заявки",
        "Unnamed: 5": "Тип контрагента",
        "Unnamed: 8": "Дата доставки",
        "Unnamed: 11": "Время доставки",
        "Unnamed: 13": "Список товаров",
        "Unnamed: 20": "Кол-во товара",
        "Unnamed: 21": "Объем заказа",
        "Unnamed: 23": "Вес заказа",
        "Unnamed: 25": "Адрес доставки",
        "Unnamed: 27": "Телефон клиента",
        "Unnamed: 29": "Способ оплаты",
        "Unnamed: 35": "Стоимость заказа, руб.",
        "Unnamed: 51": "Комментарий"
    })

    required_columns = [
        "Номер заявки",
        "Тип контрагента",
        "Дата доставки",
        "Время доставки",
        "Список товаров",
        "Кол-во товара",
        "Объем заказа",
        "Вес заказа",
        "Адрес доставки",
        "Телефон клиента",
        "Способ оплаты",
        "Стоимость заказа, руб.",
        "Комментарий",
        "Филиал"
    ]

    df = df[required_columns].copy()

    df["Номер заявки"] = df["Номер заявки"].astype("string").str.strip()
    df["Телефон клиента"] = df["Телефон клиента"].apply(clean_phone)

    # --------------------------------------------------
    # 7. Разбиваем время доставки
    # --------------------------------------------------

    time_text = (
        df["Время доставки"]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
    )

    df[["Время С", "Время ПО"]] = time_text.str.extract(
        r"[сc]\s*(\d{1,2}[:.]\d{2})\s*(?:по|до)\s*(\d{1,2}[:.]\d{2})",
        flags=re.IGNORECASE,
        expand=True
    )

    df["Время С"] = df["Время С"].str.replace(".", ":", regex=False)
    df["Время ПО"] = df["Время ПО"].str.replace(".", ":", regex=False)

    # --------------------------------------------------
    # 8. Полный обработанный файл
    # --------------------------------------------------

    full_df = df.copy()

    # --------------------------------------------------
    # 9. Подготавливаем вес
    # --------------------------------------------------

    df["Вес заказа"] = (
        df["Вес заказа"]
        .astype(str)
        .str.replace(",", ".", regex=False)
    )

    df["Вес заказа"] = pd.to_numeric(df["Вес заказа"], errors="coerce")

    # Дополнительная гарантия является услугой, а не физическим товаром.
    # Строка остаётся в полном файле и в перечне товаров, но её вес
    # не участвует в агрегировании сгруппированного и урезанного файлов.
    additional_warranty_mask = (
        df["Список товаров"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.contains(r"доп\.?\s*гарантия", case=False, na=False, regex=True)
    )

    df.loc[additional_warranty_mask, "Вес заказа"] = 0

    # --------------------------------------------------
    # 10. Сгруппированный файл
    # --------------------------------------------------

    group_columns = [
        "Дата доставки",
        "Номер заявки",
        "Филиал",
        "Адрес доставки"
    ]

    grouped_base = (
        df.groupby(
            group_columns,
            as_index=False,
            dropna=False,
            sort=False
        )
        .agg({
            "Тип контрагента": "first",
            "Время С": "first",
            "Время ПО": "first",
            "Телефон клиента": "first",
            "Способ оплаты": "first",
            "Стоимость заказа, руб.": "first",
            "Вес заказа": "sum",
            "Комментарий": "first"
        })
        .rename(columns={
            "Тип контрагента": "Тип клиента"
        })
    )

    products_df = make_products_df(df, group_columns)

    grouped_df = grouped_base.merge(
        products_df,
        on=group_columns,
        how="left"
    )

    # --------------------------------------------------
    # 11. Добавляем зону
    # --------------------------------------------------

    grouped_df["Зона"] = grouped_df["Адрес доставки"].apply(
        determine_delivery_zone
    )

    # --------------------------------------------------
    # 12. Добавляем водителей, если файл загружен
    # --------------------------------------------------

    if drivers_file is not None:
        df_vod = prepare_drivers_file(drivers_file)

        grouped_df["Номер заявки"] = grouped_df["Номер заявки"].astype("string").str.strip()
        df_vod["Номер заявки"] = df_vod["Номер заявки"].astype("string").str.strip()

        grouped_df = grouped_df.merge(
            df_vod,
            on="Номер заявки",
            how="left"
        )

    # --------------------------------------------------
    # 13. Финальный порядок столбцов сгруппированного файла
    # --------------------------------------------------

    final_columns = [
        "Дата доставки",
        "Номер заявки",
        "Филиал",
        "Тип клиента",
        "Адрес доставки",
        "Телефон клиента",
        "Время С",
        "Время ПО",
        "Способ оплаты",
        "Стоимость заказа, руб.",
        "Вес заказа",
        "Товары заявки",
        "Комментарий",
        "Зона"
    ]

    if "Водитель" in grouped_df.columns:
        final_columns.append("Водитель")

    for col in final_columns:
        if col not in grouped_df.columns:
            grouped_df[col] = pd.NA

    grouped_df = grouped_df[final_columns].copy()

    # --------------------------------------------------
    # 14. Урезанный файл
    # --------------------------------------------------

    short_df = make_short_df(grouped_df)

    # --------------------------------------------------
    # 15. Названия файлов
    # --------------------------------------------------

    original_stem = Path(original_filename).stem

    full_filename = f"Полный_{original_stem}.xlsx"

    if drivers_file is not None:
        grouped_filename = f"Водители_Зона_СГруппированный_{original_stem}.xlsx"
        short_filename = f"Урезанный_Водители_Зона_{original_stem}.xlsx"
    else:
        grouped_filename = f"Зона_СГруппированный_{original_stem}.xlsx"
        short_filename = f"Урезанный_Зона_{original_stem}.xlsx"

    return full_df, grouped_df, short_df, full_filename, grouped_filename, short_filename


def normalize_compact_text(value) -> str:
    if pd.isna(value):
        return ""

    value_text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", value_text)


def first_non_empty(series: pd.Series):
    for value in series:
        if pd.isna(value):
            continue

        if isinstance(value, str):
            cleaned = normalize_compact_text(value)

            if cleaned != "":
                return cleaned
        else:
            return value

    return pd.NA


def split_phone_values(value) -> list[str]:
    if pd.isna(value):
        return []

    value_text = str(value).replace("\xa0", " ").strip()

    if value_text == "":
        return []

    parts = re.split(r"[;\n\r]+", value_text)
    phones = []

    for part in parts:
        cleaned_phone = clean_phone(part)

        if pd.isna(cleaned_phone):
            continue

        cleaned_phone = str(cleaned_phone).strip()

        if cleaned_phone != "":
            phones.append(cleaned_phone)

    return phones


def join_unique_phones(series: pd.Series) -> str:
    phones = []
    seen = set()

    for value in series:
        for phone in split_phone_values(value):
            if phone not in seen:
                seen.add(phone)
                phones.append(phone)

    return "; ".join(phones)


def join_products_for_bitrix24(group: pd.DataFrame) -> str:
    product_lines = []

    for _, row in group.iterrows():
        product = row.get("Список товаров")
        quantity = row.get("Кол-во товара")

        if pd.isna(product):
            continue

        product_text = normalize_compact_text(product)

        if product_text == "":
            continue

        quantity_text = format_quantity(quantity)

        if quantity_text != "":
            product_lines.append(f"{product_text} - {quantity_text}шт.")
        else:
            product_lines.append(product_text)

    return "; ".join(product_lines)


def make_bitrix24_export_df(prepared_df: pd.DataFrame) -> pd.DataFrame:
    df = prepared_df.copy()

    df["Номер заявки"] = (
        df["Номер заявки"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
    )

    df = df[
        df["Номер заявки"].notna()
        & (df["Номер заявки"] != "")
    ].copy()

    df["Телефон клиента"] = df["Телефон клиента"].apply(clean_phone)

    df["Вес заказа"] = (
        df["Вес заказа"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df["Вес заказа"] = pd.to_numeric(df["Вес заказа"], errors="coerce")

    normalized_products = (
        df["Список товаров"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.lower()
    )

    delivery_service_mask = normalized_products.eq("доставка товара клиенту")
    df.loc[delivery_service_mask, "Вес заказа"] = 0

    rows = []

    for order_number, group in df.groupby(
        "Номер заявки",
        sort=False,
        dropna=False
    ):
        order_weight = group["Вес заказа"].sum(min_count=1)

        rows.append({
            "Название": order_number,
            "Номер заявки": order_number,
            "Дата доставки": first_non_empty(group["Дата доставки"]),
            "Время доставки": first_non_empty(group["Время доставки"]),
            "Список товаров": join_products_for_bitrix24(group),
            "Вес заказа, кг": order_weight,
            "Адрес доставки": first_non_empty(group["Адрес доставки"]),
            "Адрес на карте": "",
            "Телефоны клиента": join_unique_phones(group["Телефон клиента"]),
            "Способ оплаты": first_non_empty(group["Способ оплаты"]),
            "Стоимость заказа": first_non_empty(group["Стоимость заказа, руб."]),
            "Комментарий": first_non_empty(group["Комментарий"]),
            "Водитель": "",
            "Номер маршрута": "",
            "Логист": "",
        })

    result_columns = [
        "Название",
        "Номер заявки",
        "Дата доставки",
        "Время доставки",
        "Список товаров",
        "Вес заказа, кг",
        "Адрес доставки",
        "Адрес на карте",
        "Телефоны клиента",
        "Способ оплаты",
        "Стоимость заказа",
        "Комментарий",
        "Водитель",
        "Номер маршрута",
        "Логист"
    ]

    return pd.DataFrame(rows, columns=result_columns)


def process_bitrix24_file(
    main_file,
    original_filename: str = "file.xlsx"
) -> tuple[pd.DataFrame, str]:
    df = pd.read_excel(
        main_file,
        sheet_name="TDSheet"
    )

    current_filial = None
    filial_values = []

    for row_index in range(len(df)):
        first_col_value = df.iloc[row_index, 0]

        if df.shape[1] > 4:
            fifth_col_value = df.iloc[row_index, 4]
        else:
            fifth_col_value = None

        if pd.notna(first_col_value) and "Филиал" in str(first_col_value):
            if pd.notna(fifth_col_value) and str(fifth_col_value).strip() != "":
                current_filial = str(fifth_col_value).strip()

        filial_values.append(current_filial)

    df["Филиал"] = filial_values
    df = df.replace(r"^\s*$", pd.NA, regex=True)

    first_col = df.iloc[:, 0].astype(str)

    mask_delete = (
        first_col.str.contains("Филиал:", na=False)
        | first_col.str.contains("Список заявок на доставку", na=False)
        | first_col.str.contains("Доставочная организация:", na=False)
        | first_col.str.contains("Дата:", na=False)
        | first_col.str.contains("№", na=False)
    )

    mask_itogo = df.astype("string").apply(
        lambda column: column.str.contains("ИТОГО", case=False, na=False)
    ).any(axis=1)

    cols_except_filial = [
        column
        for column in df.columns
        if column != "Филиал"
    ]
    mask_empty_except_filial = df[cols_except_filial].isna().all(axis=1)

    df = df[
        ~(mask_delete | mask_itogo | mask_empty_except_filial)
    ].reset_index(drop=True)

    source_columns = [
        "Unnamed: 3",
        "Unnamed: 8",
        "Unnamed: 11",
        "Unnamed: 13",
        "Unnamed: 20",
        "Unnamed: 23",
        "Unnamed: 25",
        "Unnamed: 27",
        "Unnamed: 29",
        "Unnamed: 35",
        "Unnamed: 51"
    ]

    for column in source_columns:
        if column not in df.columns:
            df[column] = pd.NA

    df = df[source_columns].copy()

    comment_column = "Unnamed: 51"
    columns_to_fill = [
        column
        for column in df.columns
        if column != comment_column
    ]

    df[columns_to_fill] = df[columns_to_fill].replace(
        r"^\s*$",
        pd.NA,
        regex=True
    )
    df[columns_to_fill] = df[columns_to_fill].ffill()

    df = df.rename(columns={
        "Unnamed: 3": "Номер заявки",
        "Unnamed: 8": "Дата доставки",
        "Unnamed: 11": "Время доставки",
        "Unnamed: 13": "Список товаров",
        "Unnamed: 20": "Кол-во товара",
        "Unnamed: 23": "Вес заказа",
        "Unnamed: 25": "Адрес доставки",
        "Unnamed: 27": "Телефон клиента",
        "Unnamed: 29": "Способ оплаты",
        "Unnamed: 35": "Стоимость заказа, руб.",
        "Unnamed: 51": "Комментарий"
    })

    bitrix_df = make_bitrix24_export_df(df)

    original_stem = Path(original_filename).stem
    bitrix_filename = f"{original_stem}_Битрикс24.csv"

    return bitrix_df, bitrix_filename




def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    export_df = df.copy()

    for column in export_df.columns:
        if export_df[column].dtype == "object":
            export_df[column] = export_df[column].apply(
                lambda value: value.replace(";", "&")
                if isinstance(value, str)
                else value
            )

    return export_df.to_csv(
        index=False,
        sep=";",
        encoding="utf-8-sig"
    ).encode("utf-8-sig")



def parse_numeric_value(value):
    if pd.isna(value):
        return None

    value_text = str(value).replace("\xa0", " ").strip()

    if value_text == "":
        return None

    value_text = value_text.replace(" ", "").replace(",", ".")

    try:
        return float(value_text)
    except ValueError:
        return None


def format_number_for_excel(value):
    if value is None or pd.isna(value):
        return None

    number = float(value)

    if number.is_integer():
        return int(number)

    return number


def format_yamroute_time_window(value) -> str:
    value_text = normalize_compact_text(value)

    if value_text == "":
        return ""

    match = re.search(
        r"[сc]\s*(\d{1,2}[:.]\d{2})\s*(?:по|до)\s*(\d{1,2}[:.]\d{2})",
        value_text,
        flags=re.IGNORECASE
    )

    if match is None:
        return value_text

    time_from = match.group(1).replace(".", ":")
    time_to = match.group(2).replace(".", ":")

    return f"{time_from} - {time_to}"


def make_yamroute_orders_df(prepared_df: pd.DataFrame) -> pd.DataFrame:
    df = prepared_df.copy()

    df["Номер заявки"] = (
        df["Номер заявки"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
    )

    df = df[
        df["Номер заявки"].notna()
        & (df["Номер заявки"] != "")
    ].copy()

    df["Вес заказа"] = df["Вес заказа"].apply(parse_numeric_value)
    df["Кол-во товара"] = df["Кол-во товара"].apply(parse_numeric_value)
    df["Объем заказа"] = df["Объем заказа"].apply(parse_numeric_value)

    normalized_products = (
        df["Список товаров"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.lower()
    )

    delivery_service_mask = normalized_products.eq("доставка товара клиенту")

    df["Вес для ЯМаршрут"] = df["Вес заказа"]
    df.loc[delivery_service_mask, "Вес для ЯМаршрут"] = 0

    df["Количество мест для ЯМаршрут"] = df["Кол-во товара"]
    df.loc[delivery_service_mask, "Количество мест для ЯМаршрут"] = 0

    rows = []

    for order_number, group in df.groupby(
        "Номер заявки",
        sort=False,
        dropna=False
    ):
        order_weight = pd.to_numeric(
            group["Вес для ЯМаршрут"],
            errors="coerce"
        ).sum(min_count=1)

        order_units = pd.to_numeric(
            group["Количество мест для ЯМаршрут"],
            errors="coerce"
        ).sum(min_count=1)

        order_volume = pd.to_numeric(
            group["Объем заказа"],
            errors="coerce"
        ).sum(min_count=1)

        rows.append({
            "Идентификатор заказа": order_number,
            "Широта": None,
            "Долгота": None,
            "Клиент": first_non_empty(group["Тип контрагента"]),
            "Адрес": first_non_empty(group["Адрес доставки"]),
            "Временное окно": format_yamroute_time_window(
                first_non_empty(group["Время доставки"])
            ),
            "Комментарий": first_non_empty(group["Комментарий"]),
            "Жесткое окно, TRUE/FALSE": "ЛОЖЬ",
            "Время обслуживания на адрес, сек": 300,
            "Время обслуживания на заказ, сек": 300,
            "Вес, кг": format_number_for_excel(order_weight),
            "Количество мест": format_number_for_excel(order_units),
            "Объем заказа, м3": format_number_for_excel(order_volume),
            "Привязка к складам": None,
        })

    result_columns = [
        "Идентификатор заказа",
        "Широта",
        "Долгота",
        "Клиент",
        "Адрес",
        "Временное окно",
        "Комментарий",
        "Жесткое окно, TRUE/FALSE",
        "Время обслуживания на адрес, сек",
        "Время обслуживания на заказ, сек",
        "Вес, кг",
        "Количество мест",
        "Объем заказа, м3",
        "Привязка к складам",
    ]

    return pd.DataFrame(rows, columns=result_columns)



def normalize_yamroute_fixed_sheets(workbook) -> None:
    """
    Приводит постоянные листы шаблона ЯМаршрут к формату,
    который ожидает импорт:
    - ПРАВДА на листе Vehicles заменяется на текст TRUE;
    - широта и долгота на листе Depot сохраняются как числа.
    """
    vehicles_sheet = workbook["Vehicles"]

    for row in vehicles_sheet.iter_rows():
        for cell in row:
            if (
                isinstance(cell.value, str)
                and cell.value.strip().upper() == "ПРАВДА"
            ):
                cell.value = "TRUE"

    depot_sheet = workbook["Depot"]

    latitude_column = None
    longitude_column = None
    technical_header_row = None

    for row in depot_sheet.iter_rows(
        min_row=1,
        max_row=min(depot_sheet.max_row, 10)
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            normalized_value = cell.value.replace("\xa0", " ").strip().lower()

            if normalized_value == "point.lat":
                latitude_column = cell.column
                technical_header_row = cell.row

            elif normalized_value == "point.lon":
                longitude_column = cell.column
                technical_header_row = cell.row

    if latitude_column is None or longitude_column is None or technical_header_row is None:
        raise ValueError(
            "На листе Depot не найдены технические столбцы point.lat и point.lon."
        )

    for row_number in range(technical_header_row + 1, depot_sheet.max_row + 1):
        for column_number, field_name in [
            (latitude_column, "Широта склада"),
            (longitude_column, "Долгота склада"),
        ]:
            cell = depot_sheet.cell(
                row=row_number,
                column=column_number
            )

            if cell.value is None:
                continue

            if isinstance(cell.value, bool):
                raise ValueError(
                    f"{field_name} в строке {row_number} указана некорректно."
                )

            if isinstance(cell.value, (int, float)):
                cell.value = float(cell.value)
                cell.number_format = "0.000000"
                continue

            value_text = (
                str(cell.value)
                .replace("\xa0", " ")
                .replace(" ", "")
                .replace(",", ".")
                .strip()
            )

            if value_text == "":
                cell.value = None
                continue

            try:
                cell.value = float(value_text)
                cell.number_format = "0.000000"
            except ValueError as error:
                raise ValueError(
                    f"{field_name} в строке {row_number} не является числом: "
                    f"{cell.value}"
                ) from error


def process_yamroute_file(
    main_file,
    original_filename: str = "file.xlsx",
    template_path=None
) -> tuple[pd.DataFrame, bytes, str]:
    df = pd.read_excel(
        main_file,
        sheet_name="TDSheet"
    )

    current_filial = None
    filial_values = []

    for row_index in range(len(df)):
        first_col_value = df.iloc[row_index, 0]

        if df.shape[1] > 4:
            fifth_col_value = df.iloc[row_index, 4]
        else:
            fifth_col_value = None

        if pd.notna(first_col_value) and "Филиал" in str(first_col_value):
            if pd.notna(fifth_col_value) and str(fifth_col_value).strip() != "":
                current_filial = str(fifth_col_value).strip()

        filial_values.append(current_filial)

    df["Филиал"] = filial_values
    df = df.replace(r"^\s*$", pd.NA, regex=True)

    first_col = df.iloc[:, 0].astype(str)

    mask_delete = (
        first_col.str.contains("Филиал:", na=False)
        | first_col.str.contains("Список заявок на доставку", na=False)
        | first_col.str.contains("Доставочная организация:", na=False)
        | first_col.str.contains("Дата:", na=False)
        | first_col.str.contains("№", na=False)
    )

    mask_itogo = df.astype("string").apply(
        lambda column: column.str.contains("ИТОГО", case=False, na=False)
    ).any(axis=1)

    cols_except_filial = [
        column
        for column in df.columns
        if column != "Филиал"
    ]
    mask_empty_except_filial = df[cols_except_filial].isna().all(axis=1)

    df = df[
        ~(mask_delete | mask_itogo | mask_empty_except_filial)
    ].reset_index(drop=True)

    source_columns = [
        "Unnamed: 3",
        "Unnamed: 5",
        "Unnamed: 8",
        "Unnamed: 11",
        "Unnamed: 13",
        "Unnamed: 20",
        "Unnamed: 21",
        "Unnamed: 23",
        "Unnamed: 25",
        "Unnamed: 51"
    ]

    for column in source_columns:
        if column not in df.columns:
            df[column] = pd.NA

    df = df[source_columns].copy()

    comment_column = "Unnamed: 51"
    columns_to_fill = [
        column
        for column in df.columns
        if column != comment_column
    ]

    df[columns_to_fill] = df[columns_to_fill].replace(
        r"^\s*$",
        pd.NA,
        regex=True
    )
    df[columns_to_fill] = df[columns_to_fill].ffill()

    df = df.rename(columns={
        "Unnamed: 3": "Номер заявки",
        "Unnamed: 5": "Тип контрагента",
        "Unnamed: 8": "Дата доставки",
        "Unnamed: 11": "Время доставки",
        "Unnamed: 13": "Список товаров",
        "Unnamed: 20": "Кол-во товара",
        "Unnamed: 21": "Объем заказа",
        "Unnamed: 23": "Вес заказа",
        "Unnamed: 25": "Адрес доставки",
        "Unnamed: 51": "Комментарий"
    })

    orders_df = make_yamroute_orders_df(df)

    if template_path is None:
        template_path = Path(__file__).with_name("yamroute_template.xlsx")
    else:
        template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(
            "Не найден шаблон yamroute_template.xlsx. "
            "Загрузите его в корень проекта рядом с processing.py."
        )

    workbook = load_workbook(template_path)

    required_sheets = {"Orders", "Vehicles", "Depot"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)

    if missing_sheets:
        raise ValueError(
            "В шаблоне ЯМаршрут отсутствуют листы: "
            f"{', '.join(sorted(missing_sheets))}."
        )

    normalize_yamroute_fixed_sheets(workbook)

    orders_sheet = workbook["Orders"]

    for row in orders_sheet.iter_rows(
        min_row=4,
        max_row=max(orders_sheet.max_row, 4),
        min_col=1,
        max_col=26
    ):
        for cell in row:
            cell.value = None

    output_columns = [
        "Идентификатор заказа",
        "Широта",
        "Долгота",
        "Клиент",
        "Адрес",
        "Временное окно",
        "Комментарий",
        "Жесткое окно, TRUE/FALSE",
        "Время обслуживания на адрес, сек",
        "Время обслуживания на заказ, сек",
        "Вес, кг",
        "Количество мест",
        "Объем заказа, м3",
        "Привязка к складам",
    ]

    for row_number, row_values in enumerate(
        orders_df[output_columns].itertuples(index=False, name=None),
        start=4
    ):
        for column_number, value in enumerate(row_values, start=1):
            if pd.isna(value):
                value = None

            orders_sheet.cell(
                row=row_number,
                column=column_number,
                value=value
            )

    output = BytesIO()
    workbook.save(output)

    original_stem = Path(original_filename).stem
    output_filename = f"{original_stem}_ЯМаршрут.xlsx"

    return orders_df, output.getvalue(), output_filename
