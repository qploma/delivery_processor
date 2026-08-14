from io import BytesIO
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


ORDER_SOURCE_COLUMNS = {
    "order_number": "Unnamed: 3",
    "client_type": "Unnamed: 5",
    "delivery_date": "Unnamed: 8",
    "delivery_time": "Unnamed: 11",
    "product": "Unnamed: 13",
    "quantity": "Unnamed: 20",
    "volume": "Unnamed: 21",
    "weight": "Unnamed: 23",
    "address": "Unnamed: 25",
    "phone": "Unnamed: 27",
    "comment": "Unnamed: 51",
    "branch_address": "Unnamed: 18",
}


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\xa0", " ").strip()
    )


def first_non_empty(series: pd.Series):
    for value in series:
        if pd.isna(value):
            continue

        text = normalize_text(value)
        if text != "":
            return text

    return pd.NA


def parse_numeric_value(value):
    if pd.isna(value):
        return None

    text = str(value).replace("\xa0", " ").strip()
    if text == "":
        return None

    text = text.replace(" ", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def format_number(value):
    if value is None or pd.isna(value):
        return None

    number = float(value)
    if number.is_integer():
        return int(number)

    return number


def clean_phone(value):
    if pd.isna(value):
        return ""

    text = str(value).replace("\xa0", " ").strip()

    if text == "":
        return ""

    text = text.replace(" ", "")

    if text.endswith(".0"):
        text = text[:-2]

    return text


def split_phone_values(value):
    if pd.isna(value):
        return []

    text = str(value).replace("\xa0", " ").strip()

    if text == "":
        return []

    chunks = re.split(r"[;,\n\r]+", text)

    result = []

    for chunk in chunks:
        chunk = chunk.strip()

        if not chunk:
            continue

        phones_in_chunk = re.findall(
            r"(?<!\d)\+?\d{10,15}(?!\d)",
            chunk
        )

        if len(phones_in_chunk) >= 2:
            parts = phones_in_chunk
        else:
            parts = [chunk]

        for part in parts:
            phone = clean_phone(part)

            if phone and phone not in result:
                result.append(phone)

    return result


def join_phones(series):
    result = []
    seen = set()

    for value in series:
        for phone in split_phone_values(value):
            if phone not in seen:
                seen.add(phone)
                result.append(phone)

    return ", ".join(result)


def normalize_product_name(value) -> str:
    return normalize_text(value).lower()


def depot_id_from_branch_address(value):
    """
    Определяет ID склада по адресу строки «Адрес:» после заголовка филиала.

    100 — Сименса, дом 3
    200 — Выборгское шоссе, дом 503
    300 — все остальные адреса
    """
    text = normalize_product_name(value)
    text = text.replace("ё", "е")

    if (
        "сименса" in text
        and re.search(r"\bдом\s*3\b", text)
    ):
        return 100

    if (
        "выборгское" in text
        and re.search(r"\bдом\s*503\b", text)
    ):
        return 200

    return 300


def _ensure_source_columns(df):
    for column in ORDER_SOURCE_COLUMNS.values():
        if column not in df.columns:
            df[column] = pd.NA


def read_source_file(main_file):
    df = pd.read_excel(main_file, sheet_name="TDSheet")
    _ensure_source_columns(df)

    current_depot_id = 300
    depot_ids = []

    for row_index in range(len(df)):
        first_value = df.iloc[row_index, 0]

        if pd.notna(first_value) and "Филиал" in str(first_value):
            branch_address = df.iloc[row_index][
                ORDER_SOURCE_COLUMNS["branch_address"]
            ]

            if pd.notna(branch_address) and normalize_text(branch_address):
                current_depot_id = depot_id_from_branch_address(
                    branch_address
                )

        depot_ids.append(current_depot_id)

    df["__depot_id"] = depot_ids

    df = df.replace(r"^\s*$", pd.NA, regex=True)

    first_col = df.iloc[:, 0].astype(str)

    mask_delete = (
        first_col.str.contains("Филиал:", na=False)
        | first_col.str.contains("Список заявок на доставку", na=False)
        | first_col.str.contains("Доставочная организация:", na=False)
        | first_col.str.contains("Дата:", na=False)
        | first_col.str.contains("№", na=False)
    )

    mask_itogo = df.astype("string").apply(
        lambda column: column.str.contains(
            "ИТОГО",
            case=False,
            na=False
        )
    ).any(axis=1)

    selected_columns = list(ORDER_SOURCE_COLUMNS.values()) + ["__depot_id"]

    mask_empty_order_data = df[
        selected_columns
    ].drop(columns=["__depot_id"]).isna().all(axis=1)

    df = df[
        ~(mask_delete | mask_itogo | mask_empty_order_data)
    ].reset_index(drop=True)

    result = pd.DataFrame({
        "Номер заявки": df[ORDER_SOURCE_COLUMNS["order_number"]],
        "Тип контрагента": df[ORDER_SOURCE_COLUMNS["client_type"]],
        "Дата доставки": df[ORDER_SOURCE_COLUMNS["delivery_date"]],
        "Время доставки": df[ORDER_SOURCE_COLUMNS["delivery_time"]],
        "Список товаров": df[ORDER_SOURCE_COLUMNS["product"]],
        "Кол-во товара": df[ORDER_SOURCE_COLUMNS["quantity"]],
        "Объем заказа": df[ORDER_SOURCE_COLUMNS["volume"]],
        "Вес заказа": df[ORDER_SOURCE_COLUMNS["weight"]],
        "Адрес доставки": df[ORDER_SOURCE_COLUMNS["address"]],
        "Телефон клиента": df[ORDER_SOURCE_COLUMNS["phone"]],
        "Комментарий": df[ORDER_SOURCE_COLUMNS["comment"]],
        "Привязка к складам": df["__depot_id"],
    })

    fill_columns = [
        "Номер заявки",
        "Тип контрагента",
        "Дата доставки",
        "Время доставки",
        "Список товаров",
        "Кол-во товара",
        "Объем заказа",
        "Вес заказа",
        "Адрес доставки",
        "Телефон клиента",
        "Привязка к складам",
    ]

    result[fill_columns] = result[fill_columns].replace(
        r"^\s*$",
        pd.NA,
        regex=True
    )
    result[fill_columns] = result[fill_columns].ffill()

    return result


def build_orders_df(prepared_df):
    df = prepared_df.copy()

    df["Номер заявки"] = (
        df["Номер заявки"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
    )

    df = df[
        df["Номер заявки"].notna()
        & (df["Номер заявки"] != "")
    ].copy()

    df["Вес заказа"] = df["Вес заказа"].apply(parse_numeric_value)
    df["Кол-во товара"] = df["Кол-во товара"].apply(parse_numeric_value)
    df["Объем заказа"] = df["Объем заказа"].apply(parse_numeric_value)

    product_normalized = (
        df["Список товаров"]
        .astype("string")
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.lower()
    )

    delivery_mask = product_normalized.eq(
        "доставка товара клиенту"
    )

    # Услуга «Доставка товара клиенту» имеет технический вес 0,001
    # и количество 1, но в логистический вес/места не должна попадать.
    df["Вес для ЯМаршрут"] = df["Вес заказа"]
    df.loc[delivery_mask, "Вес для ЯМаршрут"] = 0

    df["Места для ЯМаршрут"] = df["Кол-во товара"]
    df.loc[delivery_mask, "Места для ЯМаршрут"] = 0

    rows = []

    for order_number, group in df.groupby(
        "Номер заявки",
        sort=False,
        dropna=False
    ):
        order_weight = pd.to_numeric(
            group["Вес для ЯМаршрут"],
            errors="coerce"
        ).sum(min_count=1)

        order_units = pd.to_numeric(
            group["Места для ЯМаршрут"],
            errors="coerce"
        ).sum(min_count=1)

        order_volume = pd.to_numeric(
            group["Объем заказа"],
            errors="coerce"
        ).sum(min_count=1)

        depot_values = pd.to_numeric(
            group["Привязка к складам"],
            errors="coerce"
        ).dropna()

        depot_id = (
            int(depot_values.iloc[0])
            if not depot_values.empty
            else 300
        )

        rows.append({
            "Идентификатор заявки": order_number,
            "Широта": None,
            "Долгота": None,
            "Клиент": first_non_empty(group["Тип контрагента"]),
            "Адрес": first_non_empty(group["Адрес доставки"]),
            "Телефон": join_phones(group["Телефон клиента"]),
            "Временное окно *": "09:00 - 23:59",
            "Жесткое окно, TRUE/FALSE": "ЛОЖЬ",
            "на адрес, сек": 600,
            "на заказ, сек": 120,
            "Вес, кг": format_number(order_weight),
            "Количество мест": format_number(order_units),
            "Требования к машине": "",
            "Комментарий": first_non_empty(group["Комментарий"]),
            "Объем заказа, м3": format_number(order_volume),
            "Привязка к складам": depot_id,
            "Штраф за факт раннего приезда, единицы": 1000,
            "Штраф за минуту раннего приезда, единицы": 17,
            "Штраф за факт опоздания, единицы": 1000,
            "Штраф за минуту опоздания, единицы": 17,
            "Тип заказа": "delivery",
            "Место назначения": "",
            "Тип совместимости с другими заказами": "",
            "Пресет": "public_minimize_mileage",
        })

    return pd.DataFrame(rows)


def _clear_orders_data(orders_sheet):
    # Шаблон имеет заголовки в строках 1–3, данные начинаются с 4.
    # Очищаем старую тестовую заявку и все остальные строки данных.
    max_row = max(orders_sheet.max_row, 4)
    max_col = max(orders_sheet.max_column, 26)

    for row in orders_sheet.iter_rows(
        min_row=4,
        max_row=max_row,
        min_col=1,
        max_col=max_col
    ):
        for cell in row:
            cell.value = None


def _find_template_columns(orders_sheet):
    """
    Берём технические имена из строки 3 шаблона, чтобы не зависеть
    от визуальных заголовков и объединённых ячеек.
    """
    columns = {}

    for column_number in range(1, orders_sheet.max_column + 1):
        value = orders_sheet.cell(
            row=3,
            column=column_number
        ).value

        if isinstance(value, str) and value.strip():
            columns[value.strip()] = column_number

    return columns



def _clear_depot_coordinates(depot_sheet):
    latitude_column = None
    longitude_column = None
    header_row = None

    for row in depot_sheet.iter_rows(
        min_row=1,
        max_row=min(depot_sheet.max_row, 10)
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            value = (
                cell.value
                .replace("\xa0", " ")
                .strip()
                .lower()
            )

            if value == "point.lat":
                latitude_column = cell.column
                header_row = cell.row

            elif value == "point.lon":
                longitude_column = cell.column
                header_row = cell.row

    if (
        latitude_column is None
        or longitude_column is None
        or header_row is None
    ):
        raise ValueError(
            "На листе Depot не найдены point.lat и point.lon."
        )

    for row_number in range(
        header_row + 1,
        depot_sheet.max_row + 1
    ):
        depot_sheet.cell(
            row=row_number,
            column=latitude_column
        ).value = None

        depot_sheet.cell(
            row=row_number,
            column=longitude_column
        ).value = None

def process_yamroute2_file(
    main_file,
    original_filename="file.xlsx",
    template_path=None
):
    prepared_df = read_source_file(main_file)
    orders_df = build_orders_df(prepared_df)

    if template_path is None:
        template_path = Path(__file__).with_name(
            "yamroute2_template.xlsx"
        )
    else:
        template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(
            "Не найден шаблон yamroute2_template.xlsx. "
            "Загрузите его в корень проекта рядом с yamroute2.py."
        )

    workbook = load_workbook(template_path)

    required_sheets = {"Orders", "Depot"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)

    if missing_sheets:
        raise ValueError(
            "В шаблоне 2ЯМаршрут отсутствуют листы: "
            f"{', '.join(sorted(missing_sheets))}."
        )

    depot_sheet = workbook["Depot"]
    _clear_depot_coordinates(depot_sheet)

    orders_sheet = workbook["Orders"]
    _clear_orders_data(orders_sheet)

    template_columns = _find_template_columns(orders_sheet)

    required_technical_columns = [
        "id",
        "point.lat",
        "point.lon",
        "title",
        "address",
        "phone",
        "time_window",
        "hard_window",
        "shared_service_duration_s",
        "service_duration_s",
        "shipment_size.weight_kg",
        "shipment_size.units",
        "required_tags",
        "comments",
        "shipment_size.volume_cbm",
        "depot_id",
        "penalty.early.fixed",
        "penalty.early.minute",
        "penalty.late.fixed",
        "penalty.late.minute",
        "type",
        "delivery_to",
        "load_types",
        "preset_id",
    ]

    missing_columns = [
        column
        for column in required_technical_columns
        if column not in template_columns
    ]

    if missing_columns:
        raise ValueError(
            "В листе Orders не найдены технические столбцы: "
            f"{', '.join(missing_columns)}."
        )

    for row_number, row in enumerate(
        orders_df.itertuples(index=False),
        start=4
    ):
        values = {
            "id": row[0],
            "point.lat": None,
            "point.lon": None,
            "title": row[3],
            "address": row[4],
            "phone": row[5],
            "time_window": row[6],
            "hard_window": row[7],
            "shared_service_duration_s": row[8],
            "service_duration_s": row[9],
            "shipment_size.weight_kg": row[10],
            "shipment_size.units": row[11],
            "required_tags": row[12],
            "comments": row[13],
            "shipment_size.volume_cbm": row[14],
            "depot_id": row[15],
            "penalty.early.fixed": row[16],
            "penalty.early.minute": row[17],
            "penalty.late.fixed": row[18],
            "penalty.late.minute": row[19],
            "type": row[20],
            "delivery_to": row[21],
            "load_types": row[22],
            "preset_id": row[23],
        }

        for technical_name, value in values.items():
            if pd.isna(value):
                value = None

            orders_sheet.cell(
                row=row_number,
                column=template_columns[technical_name],
                value=value
            )

    output = BytesIO()
    workbook.save(output)

    original_stem = Path(original_filename).stem
    output_filename = f"{original_stem}_2ЯМаршрут.xlsx"

    return orders_df, output.getvalue(), output_filename
